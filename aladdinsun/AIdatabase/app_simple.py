"""
TDSQL 部署资源预测系统 - 完整版
整合：部署预测、模型库管理、自主训练
"""

# 修复OpenBLAS线程资源问题 - 必须在导入任何科学计算库之前设置
import os
os.environ['OPENBLAS_NUM_THREADS'] = '4'
os.environ['OMP_NUM_THREADS'] = '4'
os.environ['MKL_NUM_THREADS'] = '4'

from flask import Flask, render_template, request, jsonify
import json
from werkzeug.utils import secure_filename
from datetime import datetime

app = Flask(__name__)

# 配置
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'xlsx', 'xls', 'pdf', 'json', 'txt'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024

# 创建必要目录
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs('model_libraries', exist_ok=True)
os.makedirs('training_data', exist_ok=True)

print("\n" + "=" * 60)
print("🚀 TDSQL 部署资源预测系统 v4.2 (完整版)")
print("=" * 60)

# 延迟导入模块
predictor = None
library_manager = None
training_system = None
model = None

def get_predictor():
    """延迟加载预测器"""
    global predictor
    if predictor is None:
        print("📦 正在加载预测引擎...")
        from deployment_predictor import DeploymentResourcePredictor
        predictor = DeploymentResourcePredictor()
        print("✅ 预测引擎加载完成")
    return predictor

def get_library_manager():
    """延迟加载模型库管理器"""
    global library_manager
    if library_manager is None:
        print("📦 正在加载模型库管理器...")
        from model_library_manager import ModelLibraryManager
        library_manager = ModelLibraryManager()
        print("✅ 模型库管理器加载完成")
    return library_manager

def get_training_system():
    """延迟加载训练系统"""
    global training_system, model
    if training_system is None:
        print("📦 正在加载训练系统...")
        from model import TDSQLArchitecturePredictor
        from training_system import TrainingSystem
        model = TDSQLArchitecturePredictor()
        training_system = TrainingSystem(model)
        print("✅ 训练系统加载完成")
    return training_system

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ==================== 页面路由 ====================

@app.route('/')
def index():
    """主页 - 导航页面"""
    return render_template('navigation.html')

@app.route('/nav')
def navigation():
    """导航页面（别名）"""
    return render_template('navigation.html')

@app.route('/predict')
def predict_page():
    """部署资源预测页面（新版）"""
    return render_template('predict_v2.html')

@app.route('/predict_old')
def predict_page_old():
    """部署资源预测页面（旧版）"""
    return render_template('index.html')

@app.route('/model_library')
def model_library():
    """模型库管理页面"""
    return render_template('model_library.html')

@app.route('/learning')
def learning():
    """学习系统页面"""
    return render_template('index_learning.html')

@app.route('/test_predict')
def test_predict():
    """测试预测接口页面"""
    return render_template('test_simple.html')

@app.route('/test_debug')
def test_debug():
    """调试测试页面"""
    return render_template('test_simple.html')

# ==================== API路由 ====================

@app.route('/api/health', methods=['GET'])
def health_check():
    """健康检查"""
    return jsonify({
        'status': 'ok',
        'version': '4.0',
        'message': 'TDSQL部署资源预测系统运行正常'
    })

@app.route('/api/predict', methods=['POST'])
def predict():
    """部署资源预测API"""
    try:
        raw = request.get_json() or {}
        
        # 调试：打印收到的原始数据
        print("\n" + "=" * 60)
        print("📥 收到的请求参数:")
        print(f"  enable_xinchuan: {raw.get('enable_xinchuan')}")
        print(f"  xinchuan_mode: {raw.get('xinchuan_mode')}")
        print(f"  全部参数: {raw}")
        print("=" * 60 + "\n")
        
        # 统一字段映射（兼容普通版/专业版表单字段）
        data = {}
        
        # 数据规模
        data['data_volume'] = (
            raw.get('current_data_size_gb')
            or raw.get('total_data_size_gb')
            or raw.get('future_data_size_gb')
            or raw.get('data_volume')
            or 100
        )
        
        # 性能参数
        data['qps'] = raw.get('qps') or raw.get('normal_qps') or raw.get('peak_qps') or 1000
        data['tps'] = raw.get('tps') or raw.get('normal_tps') or raw.get('peak_tps') or int(data['qps'] * 0.3)
        
        # 并发/连接
        data['concurrent_users'] = (
            raw.get('concurrent_users')
            or raw.get('concurrent_connections')
            or raw.get('avg_concurrent_connections')
            or 100
        )
        
        # 行业/HA
        data['industry'] = raw.get('industry') or raw.get('industry_type') or 'general'
        need_ha = raw.get('need_high_availability') is True
        data['ha_level'] = 'high' if need_ha else 'standard'
        
        # 数据增长率（转为比例）
        growth = raw.get('data_growth_rate')
        if isinstance(growth, (int, float)) and growth > 1:
            data['data_growth_rate'] = float(growth) / 100.0
        else:
            data['data_growth_rate'] = growth if isinstance(growth, (int, float)) else 0.3
        
        # 其他可选需求
        data['need_disaster_recovery'] = bool(raw.get('need_disaster_recovery'))
        data['need_read_write_split'] = bool(raw.get('need_read_write_split'))
        
        # 信创模式参数
        enable_xinchuan = raw.get('enable_xinchuan', False)  # 默认关闭，需要用户主动勾选
        xinchuan_mode = raw.get('xinchuan_mode', 'standard')  # 默认标准信创
        
        # 无论是否启用信创模式，都使用新版预测器（确保生成完整的设备清单）
        from deployment_predictor_xinchuan import DeploymentResourcePredictorXinChuan
        
        # 转换数据格式（统一定义）
        common_data = {
            'data_size_gb': data['data_volume'] * 1024,  # 转换为GB
            'transactions_per_day': data['tps'] * 86400,
            'max_connections': data['concurrent_users'],
            'business_type': 'OLTP',
            'high_availability': need_ha,
            'disaster_recovery': data['need_disaster_recovery']
        }
        
        # 如果启用信创模式，生成传统方案和信创方案的完整对比
        if enable_xinchuan:
            
            # 生成传统方案(使用国外品牌设备) - 独立架构设计
            traditional_predictor = DeploymentResourcePredictorXinChuan(xinchuan_mode='off')
            traditional_result = traditional_predictor.predict(common_data)
            
            # 生成信创方案(使用国产设备) - 独立架构设计
            xc_predictor = DeploymentResourcePredictorXinChuan(xinchuan_mode=xinchuan_mode)
            xc_result = xc_predictor.predict(common_data)
            
            # 计算真实的成本差异
            traditional_cost = traditional_result.get('cost_breakdown', {}).get('total_initial_cost', 0)
            xinchuan_cost = xc_result.get('cost_breakdown', {}).get('total_initial_cost', 0)
            cost_savings = traditional_cost - xinchuan_cost
            savings_percent = (cost_savings / traditional_cost * 100) if traditional_cost > 0 else 0
            
            # 使用传统方案作为基础结果（确保前端显示一致）
            result = traditional_result.copy()
            
            # 先获取成本明细（避免变量未定义错误）
            traditional_cost_breakdown = traditional_result.get('cost_breakdown', {})
            traditional_equipment = traditional_result.get('equipment_list', [])
            traditional_architecture = traditional_result.get('architecture', {})
            
            # 适配前端字段名（兼容旧版显示）
            result['cost'] = {
                'initial_investment': traditional_cost,  # 映射到旧字段
                'three_year_tco': traditional_cost * 1.5,  # 估算3年TCO
                'total_hardware': traditional_cost_breakdown.get('hardware_cost', 0),
                'total_software': traditional_cost_breakdown.get('software_cost', 0),
                'annual_operating': traditional_cost * 0.15  # 估算年运营成本
            }
            
            # 调试：打印设备清单数据（非信创模式）
            print(f"\n🔍 调试(非信创模式) - traditional_equipment 总数: {len(traditional_equipment)}")
            if traditional_equipment:
                print(f"🔍 调试(非信创模式) - 第一个设备示例: {traditional_equipment[0]}")
            else:
                print("⚠️  调试(非信创模式) - traditional_equipment 为空列表！")
            
            # 适配设备清单字段（按类别分组）
            servers = [item for item in traditional_equipment if item.get('category') in ['数据库服务器', '代理服务器', '监控服务器']]
            network_devices = [item for item in traditional_equipment if item.get('category') in ['核心交换机', '接入交换机', '安全防火墙']]
            storage_devices = [item for item in traditional_equipment if item.get('category') == '存储设备']
            infrastructure_items = traditional_cost_breakdown.get('infrastructure_items', [])
            
            print(f"🔍 调试(非信创模式) - 分类后: servers={len(servers)}, network={len(network_devices)}, storage={len(storage_devices)}, infrastructure={len(infrastructure_items)}")
            
            result['equipment'] = {
                'servers': servers,
                'network_devices': network_devices,
                'storage': storage_devices,  # ✅ 改为 storage（前端期待）
                'infrastructure': infrastructure_items  # ✅ 添加基础设施清单
            }
            result['equipment_list'] = traditional_equipment  # 新版字段

            
            # 适配架构字段
            result['architecture'] = {
                'type': 'cluster',
                'topology': {
                    'db_nodes': traditional_architecture.get('database_nodes', 0),
                    'proxy_nodes': traditional_architecture.get('proxy_nodes', 0),
                    'monitor_nodes': traditional_architecture.get('monitoring_nodes', 0),
                    'shard_count': 0,
                    'replica_count': 3
                }
            }
            result['cost_breakdown'] = {
                'summary': result['cost'],
                'breakdown': {  # ✅ 添加 breakdown 包裹层
                    'hardware': {
                        'servers': sum(item.get('total_price', 0) for item in servers),
                        'network_devices': sum(item.get('total_price', 0) for item in network_devices),
                        'storage': sum(item.get('total_price', 0) for item in storage_devices),
                        'infrastructure': traditional_cost_breakdown.get('infrastructure_cost', 0)
                    },
                    'software': {
                        'tdsql_license': 0,  # 信创模式免费
                        'os_license': sum(item.get('total', 0) for item in traditional_cost_breakdown.get('software_items', []) if 'OS' in item.get('name', '') or 'Red Hat' in item.get('name', '')),
                        'monitoring': 0,
                        'backup': 0
                    },
                    'services': {
                        'deployment': sum(item.get('total_price', 0) for item in infrastructure_items if '实施' in item.get('category', '')),
                        'training': sum(item.get('total_price', 0) for item in infrastructure_items if '培训' in item.get('category', ''))
                    }
                },
                # 同时保留顶层字段（向后兼容）
                'hardware': {
                    'servers': sum(item.get('total_price', 0) for item in servers),
                    'network_devices': sum(item.get('total_price', 0) for item in network_devices),
                    'storage': sum(item.get('total_price', 0) for item in storage_devices),
                    'infrastructure': traditional_cost_breakdown.get('infrastructure_cost', 0)
                },
                'software': {
                    'tdsql_license': 0,
                    'os_license': sum(item.get('total', 0) for item in traditional_cost_breakdown.get('software_items', []) if 'OS' in item.get('name', '') or 'Red Hat' in item.get('name', '')),
                    'monitoring': 0,
                    'backup': 0
                },
                'services': {
                    'deployment': sum(item.get('total_price', 0) for item in infrastructure_items if '实施' in item.get('category', '')),
                    'training': sum(item.get('total_price', 0) for item in infrastructure_items if '培训' in item.get('category', ''))
                },
                'software_items': traditional_cost_breakdown.get('software_items', []),
                'annual_operating': {}
            }
            
            # 添加完整的对比信息
            result['xinchuan_enabled'] = True
            result['xinchuan_mode'] = xinchuan_mode
            result['traditional_solution'] = traditional_result  # 传统方案(独立架构)
            result['xinchuan_solution'] = xc_result  # 信创方案(独立架构)
            result['xinchuan_info'] = xc_result.get('xinchuan_info', {})
            result['cost_comparison'] = {
                'traditional_cost': traditional_cost,
                'xinchuan_cost': xinchuan_cost,
                'cost_savings': cost_savings,
                'savings_percent': round(savings_percent, 1),
                'note': f'使用信创方案相比传统方案节约 ¥{cost_savings:,.0f} ({savings_percent:.1f}%)'
            }
        else:
            # 不启用信创模式，只生成传统方案（国外品牌）
            traditional_predictor = DeploymentResourcePredictorXinChuan(xinchuan_mode='off')
            traditional_result = traditional_predictor.predict(common_data)
            
            # 获取成本和设备信息
            traditional_cost_breakdown = traditional_result.get('cost_breakdown', {})
            traditional_equipment = traditional_result.get('equipment_list', [])
            traditional_architecture = traditional_result.get('architecture', {})
            traditional_cost = traditional_cost_breakdown.get('total_initial_cost', 0)
            
            # 使用传统方案作为结果
            result = traditional_result.copy()
            
            # 适配前端字段名（兼容旧版显示）
            result['cost'] = {
                'initial_investment': traditional_cost,
                'three_year_tco': traditional_cost * 1.5,
                'total_hardware': traditional_cost_breakdown.get('hardware_cost', 0),
                'total_software': traditional_cost_breakdown.get('software_cost', 0),
                'annual_operating': traditional_cost * 0.15
            }
            
            # 调试：打印设备清单数据（非信创模式）
            print(f"\n🔍 调试(非信创模式) - traditional_equipment 总数: {len(traditional_equipment)}")
            if traditional_equipment:
                print(f"🔍 调试(非信创模式) - 第一个设备示例: {traditional_equipment[0]}")
            else:
                print("⚠️  调试(非信创模式) - traditional_equipment 为空列表！")
            
            # 适配设备清单字段（按类别分组）
            servers = [item for item in traditional_equipment if item.get('category') in ['数据库服务器', '代理服务器', '监控服务器']]
            network_devices = [item for item in traditional_equipment if item.get('category') in ['核心交换机', '接入交换机', '安全防火墙']]
            storage_devices = [item for item in traditional_equipment if item.get('category') == '存储设备']
            infrastructure_items = traditional_cost_breakdown.get('infrastructure_items', [])
            
            print(f"🔍 调试(非信创模式) - 分类后: servers={len(servers)}, network={len(network_devices)}, storage={len(storage_devices)}, infrastructure={len(infrastructure_items)}")
            
            result['equipment'] = {
                'servers': servers,
                'network_devices': network_devices,
                'storage': storage_devices,  # ✅ 改为 storage（前端期待）
                'infrastructure': infrastructure_items  # ✅ 添加基础设施清单
            }
            result['equipment_list'] = traditional_equipment

            
            # 适配架构字段
            result['architecture'] = {
                'type': 'cluster',
                'topology': {
                    'db_nodes': traditional_architecture.get('database_nodes', 0),
                    'proxy_nodes': traditional_architecture.get('proxy_nodes', 0),
                    'monitor_nodes': traditional_architecture.get('monitoring_nodes', 0),
                    'shard_count': 0,
                    'replica_count': 3
                }
            }
            
            result['cost_breakdown'] = {
                'summary': result['cost'],
                'breakdown': {  # ✅ 添加 breakdown 包裹层
                    'hardware': {
                        'servers': sum(item.get('total_price', 0) for item in servers),
                        'network_devices': sum(item.get('total_price', 0) for item in network_devices),
                        'storage': sum(item.get('total_price', 0) for item in storage_devices),
                        'infrastructure': traditional_cost_breakdown.get('infrastructure_cost', 0)
                    },
                    'software': {
                        'tdsql_license': 0,
                        'os_license': sum(item.get('total', 0) for item in traditional_cost_breakdown.get('software_items', []) if 'OS' in item.get('name', '') or 'Red Hat' in item.get('name', '')),
                        'monitoring': 0,
                        'backup': 0
                    },
                    'services': {
                        'deployment': sum(item.get('total_price', 0) for item in infrastructure_items if '实施' in item.get('category', '')),
                        'training': sum(item.get('total_price', 0) for item in infrastructure_items if '培训' in item.get('category', ''))
                    }
                },
                # 同时保留顶层字段（向后兼容）
                'hardware': {
                    'servers': sum(item.get('total_price', 0) for item in servers),
                    'network_devices': sum(item.get('total_price', 0) for item in network_devices),
                    'storage': sum(item.get('total_price', 0) for item in storage_devices),
                    'infrastructure': traditional_cost_breakdown.get('infrastructure_cost', 0)
                },
                'software': {
                    'tdsql_license': 0,
                    'os_license': sum(item.get('total', 0) for item in traditional_cost_breakdown.get('software_items', []) if 'OS' in item.get('name', '') or 'Red Hat' in item.get('name', '')),
                    'monitoring': 0,
                    'backup': 0
                },
                'services': {
                    'deployment': sum(item.get('total_price', 0) for item in infrastructure_items if '实施' in item.get('category', '')),
                    'training': sum(item.get('total_price', 0) for item in infrastructure_items if '培训' in item.get('category', ''))
                },
                'software_items': traditional_cost_breakdown.get('software_items', []),
                'annual_operating': {}
            }
            
            # 标记未启用信创模式
            result['xinchuan_enabled'] = False
        
        return jsonify({
            'success': True,
            'data': result
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """文件上传API"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '没有文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '文件名为空'}), 400
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{timestamp}_{filename}"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            # 获取预测器并解析文件
            pred = get_predictor()
            params = pred.parse_file(filepath)
            
            return jsonify({
                'success': True,
                'filename': filename,
                'filepath': filepath,
                'params': params
            })
        else:
            return jsonify({'success': False, 'error': '不支持的文件类型'}), 400
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/clear_uploads', methods=['POST'])
def clear_uploads():
    """清除上传的文件"""
    try:
        import shutil
        
        # 清空uploads目录
        if os.path.exists(UPLOAD_FOLDER):
            for filename in os.listdir(UPLOAD_FOLDER):
                file_path = os.path.join(UPLOAD_FOLDER, filename)
                try:
                    if os.path.isfile(file_path):
                        os.unlink(file_path)
                except Exception as e:
                    print(f"删除文件失败: {file_path}, 错误: {e}")
        
        return jsonify({
            'success': True,
            'message': '已清除所有上传文件'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== 模型库管理 API ====================

@app.route('/api/model_libraries', methods=['GET'])
def get_model_libraries():
    """获取可用的模型库列表"""
    try:
        manager = get_library_manager()
        libraries = manager.list_available_libraries()
        return jsonify({
            'success': True,
            'libraries': libraries
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/model_libraries/<library_id>', methods=['GET'])
def get_model_library(library_id):
    """获取指定模型库详情"""
    try:
        manager = get_library_manager()
        library = manager.get_library_info(library_id)
        return jsonify({
            'success': True,
            'library': library
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/model_libraries/<library_id>/download', methods=['POST'])
def download_model_library(library_id):
    """下载模型库"""
    try:
        manager = get_library_manager()
        result = manager.download_library(library_id)
        return jsonify({
            'success': True,
            'message': f'模型库 {library_id} 下载成功',
            'result': result
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/model_libraries/installed', methods=['GET'])
def get_installed_libraries():
    """获取已安装的模型库"""
    try:
        manager = get_library_manager()
        installed = manager.list_installed_libraries()
        return jsonify({
            'success': True,
            'libraries': installed
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/model_libraries/<library_id>/activate', methods=['POST'])
def activate_library(library_id):
    """激活模型库"""
    try:
        manager = get_library_manager()
        manager.activate_library(library_id)
        return jsonify({
            'success': True,
            'message': f'模型库 {library_id} 已激活'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/custom_library', methods=['POST'])
def create_custom_library():
    """创建自定义模型库"""
    try:
        data = request.get_json()
        manager = get_library_manager()
        result = manager.create_custom_library(
            name=data.get('name'),
            description=data.get('description'),
            industry=data.get('industry'),
            cases=data.get('cases', [])
        )
        return jsonify({
            'success': True,
            'library': result
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== 训练系统 API ====================

@app.route('/api/training/cases', methods=['GET'])
def get_training_cases():
    """获取训练案例列表"""
    try:
        trainer = get_training_system()
        cases = trainer.list_cases()
        return jsonify({
            'success': True,
            'cases': cases
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/training/cases', methods=['POST'])
def add_training_case():
    """添加训练案例"""
    try:
        data = request.get_json()
        trainer = get_training_system()
        
        case_id = trainer.add_case(
            input_data=data.get('input'),
            output_data=data.get('output'),
            feedback=data.get('feedback')
        )
        
        return jsonify({
            'success': True,
            'case_id': case_id,
            'message': '训练案例已添加'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/training/train', methods=['POST'])
def train_model():
    """训练模型"""
    try:
        data = request.get_json()
        trainer = get_training_system()
        
        result = trainer.train(
            epochs=data.get('epochs', 10),
            batch_size=data.get('batch_size', 32)
        )
        
        return jsonify({
            'success': True,
            'result': result
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/training/history', methods=['GET'])
def get_training_history():
    """获取训练历史"""
    try:
        trainer = get_training_system()
        history = trainer.get_history()
        return jsonify({
            'success': True,
            'history': history
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/training/evaluate', methods=['POST'])
def evaluate_model():
    """评估模型"""
    try:
        data = request.get_json()
        trainer = get_training_system()
        
        result = trainer.evaluate(data.get('test_data'))
        
        return jsonify({
            'success': True,
            'result': result
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/download_template/<version>')
def download_template(version):
    """下载Excel模板"""
    try:
        from flask import send_file
        if version == 'basic':
            filename = '部署资源预测模板-普通版.xlsx'
        else:
            filename = '部署资源预测模板-专业版.xlsx'
        
        filepath = os.path.join('templates', filename)
        return send_file(filepath, as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/parse_excel', methods=['POST'])
def parse_excel():
    """解析上传的Excel文件"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '没有文件上传'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '文件名为空'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': '只支持Excel文件'}), 400
        
        # 保存文件
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # 解析Excel
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        ws = wb.active
        
        data = {}
        
        # 读取数据（从第5行开始，跳过标题和表头）
        for row in range(5, ws.max_row + 1):
            param_name = ws.cell(row, 1).value
            param_value = ws.cell(row, 2).value
            
            if param_name and param_value:
                # 转换参数名为字段名
                field_mapping = {
                    '数据规模 (GB)': 'total_data_size_gb',
                    '当前数据规模 (GB)': 'current_data_size_gb',
                    '表数量': 'table_count',
                    'QPS (每秒查询数)': 'qps',
                    '日常QPS': 'normal_qps',
                    'TPS (每秒事务数)': 'tps',
                    '日常TPS': 'normal_tps',
                    '并发连接数': 'concurrent_connections',
                    '平均并发连接数': 'avg_concurrent_connections',
                    '需要高可用': 'need_high_availability',
                    '需要灾备': 'need_disaster_recovery',
                    '需要异地灾备': 'need_disaster_recovery',
                    '需要读写分离': 'need_read_write_split',
                    '数据增长率 (%/年)': 'data_growth_rate',
                }
                
                field_name = field_mapping.get(param_name)
                if field_name:
                    # 处理布尔值
                    if isinstance(param_value, str) and param_value.upper() in ['TRUE', 'FALSE']:
                        data[field_name] = param_value.upper() == 'TRUE'
                    else:
                        data[field_name] = param_value
        
        # 删除临时文件
        os.remove(filepath)
        
        return jsonify({
            'success': True,
            'data': data
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    print("\n" + "=" * 60)
    print("🚀 TDSQL 部署资源预测系统 v4.3")
    print("=" * 60)
    print(f"📍 主页面: http://127.0.0.1:18080")
    print(f"📍 导航页面: http://127.0.0.1:18080/nav")
    print(f"📍 部署预测: http://127.0.0.1:18080/predict (新版)")
    print(f"📍 模型库管理: http://127.0.0.1:18080/model_library")
    print(f"📍 学习系统: http://127.0.0.1:18080/learning")
    print("=" * 60)
    print("✨ 功能模块:")
    print("  ✅ 部署资源预测 - 普通版/专业版双模式")
    print("  ✅ Excel模板 - 下载模板，上传自动填充")
    print("  ✅ 模型库管理 - 8个预置模型库")
    print("  ✅ 自主训练 - 从实际案例中学习优化")
    print("  ✅ 文件上传 - 支持JSON/Excel/图片/PDF")
    print("=" * 60)
    print()
    
    app.run(
        host='0.0.0.0',
        port=18080,
        debug=False,
        threaded=True,
        use_reloader=False
    )
