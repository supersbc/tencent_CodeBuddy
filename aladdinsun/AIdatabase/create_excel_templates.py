#!/usr/bin/env python3
"""
创建部署资源预测的Excel模板
- 普通版模板
- 专业版模板
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_basic_template():
    """创建普通版模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "部署资源预测-普通版"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    
    # 标题样式
    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True, size=11)
    
    # 边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 标题
    ws['A1'] = 'TDSQL 部署资源预测 - 普通版'
    ws['A1'].font = Font(bold=True, size=14, color="4472C4")
    ws.merge_cells('A1:C1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 说明
    ws['A2'] = '请填写以下参数，系统将自动预测所需的部署资源'
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:C2')
    
    # 表头
    row = 4
    ws[f'A{row}'] = '参数名称'
    ws[f'B{row}'] = '参数值'
    ws[f'C{row}'] = '说明'
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].fill = header_fill
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].border = thin_border
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 数据行
    data = [
        ('数据规模 (GB)', '', '预计存储的总数据量，单位GB'),
        ('表数量', '', '数据库中的表总数'),
        ('QPS (每秒查询数)', '', '系统每秒处理的查询请求数'),
        ('TPS (每秒事务数)', '', '系统每秒处理的事务数'),
        ('并发连接数', '', '同时连接到数据库的最大连接数'),
        ('需要高可用', 'TRUE/FALSE', '是否需要高可用架构 (填写TRUE或FALSE)'),
        ('需要灾备', 'TRUE/FALSE', '是否需要异地灾备 (填写TRUE或FALSE)'),
        ('需要读写分离', 'TRUE/FALSE', '是否需要读写分离 (填写TRUE或FALSE)'),
        ('数据增长率 (%/年)', '', '预计每年数据增长的百分比'),
    ]
    
    row = 5
    for param_name, default_value, description in data:
        ws[f'A{row}'] = param_name
        ws[f'B{row}'] = default_value
        ws[f'C{row}'] = description
        
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        ws[f'C{row}'].border = thin_border
        
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        row += 1
    
    # 示例数据
    row += 1
    ws[f'A{row}'] = '示例数据'
    ws[f'A{row}'].font = Font(bold=True, size=11, color="4472C4")
    ws.merge_cells(f'A{row}:C{row}')
    
    row += 1
    example_data = [
        ('数据规模 (GB)', '500', '中型企业应用'),
        ('表数量', '300', ''),
        ('QPS (每秒查询数)', '5000', ''),
        ('TPS (每秒事务数)', '2000', ''),
        ('并发连接数', '500', ''),
        ('需要高可用', 'TRUE', ''),
        ('需要灾备', 'FALSE', ''),
        ('需要读写分离', 'TRUE', ''),
        ('数据增长率 (%/年)', '20', ''),
    ]
    
    for param_name, value, note in example_data:
        ws[f'A{row}'] = param_name
        ws[f'B{row}'] = value
        ws[f'C{row}'] = note
        ws[f'A{row}'].font = Font(italic=True, color="666666")
        ws[f'B{row}'].font = Font(italic=True, color="666666")
        ws[f'C{row}'].font = Font(italic=True, color="666666")
        row += 1
    
    wb.save('templates/部署资源预测模板-普通版.xlsx')
    print("✅ 普通版模板创建成功: templates/部署资源预测模板-普通版.xlsx")

def create_professional_template():
    """创建专业版模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "部署资源预测-专业版"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 45
    
    # 样式
    title_font = Font(bold=True, size=14, color="4472C4")
    section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True, size=10)
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 标题
    ws['A1'] = 'TDSQL 部署资源预测 - 专业版'
    ws['A1'].font = title_font
    ws.merge_cells('A1:C1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A2'] = '专业版提供更详细的参数配置，适合复杂场景和精确预测'
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells('A2:C2')
    
    row = 4
    
    # 第一部分：基础信息
    ws[f'A{row}'] = '一、基础信息'
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    basic_params = [
        ('项目名称', '', '项目或系统的名称'),
        ('行业类型', '', '如：金融、电商、游戏、互联网等'),
        ('业务场景', '', '如：核心交易、数据分析、内容管理等'),
        ('预计上线时间', '', '格式：YYYY-MM-DD'),
        ('预算范围 (万元)', '', '可接受的预算范围'),
    ]
    
    for param_name, default_value, description in basic_params:
        ws[f'A{row}'] = param_name
        ws[f'B{row}'] = default_value
        ws[f'C{row}'] = description
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        ws[f'C{row}'].border = thin_border
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].fill = input_fill
        row += 1
    
    row += 1
    
    # 第二部分：数据规模
    ws[f'A{row}'] = '二、数据规模'
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    data_params = [
        ('当前数据规模 (GB)', '', '现有数据量'),
        ('预计3年后数据规模 (GB)', '', '3年后预计的数据量'),
        ('数据库数量', '', '需要创建的数据库数量'),
        ('表数量', '', '预计的表总数'),
        ('单表最大记录数', '', '最大表的记录数'),
        ('平均行大小 (字节)', '', '每行数据的平均大小'),
        ('数据增长率 (%/年)', '', '年度数据增长百分比'),
        ('数据保留期限 (年)', '', '数据需要保留的年限'),
    ]
    
    for param_name, default_value, description in data_params:
        ws[f'A{row}'] = param_name
        ws[f'B{row}'] = default_value
        ws[f'C{row}'] = description
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        ws[f'C{row}'].border = thin_border
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].fill = input_fill
        row += 1
    
    row += 1
    
    # 第三部分：性能需求
    ws[f'A{row}'] = '三、性能需求'
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    performance_params = [
        ('日常QPS', '', '日常每秒查询数'),
        ('峰值QPS', '', '高峰期每秒查询数'),
        ('日常TPS', '', '日常每秒事务数'),
        ('峰值TPS', '', '高峰期每秒事务数'),
        ('平均并发连接数', '', '平均同时连接数'),
        ('峰值并发连接数', '', '峰值同时连接数'),
        ('平均响应时间要求 (ms)', '', '期望的平均响应时间'),
        ('99分位响应时间要求 (ms)', '', '99%请求的响应时间要求'),
        ('读写比例', '', '读操作:写操作，如 7:3'),
    ]
    
    for param_name, default_value, description in performance_params:
        ws[f'A{row}'] = param_name
        ws[f'B{row}'] = default_value
        ws[f'C{row}'] = description
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        ws[f'C{row}'].border = thin_border
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].fill = input_fill
        row += 1
    
    row += 1
    
    # 第四部分：高可用与灾备
    ws[f'A{row}'] = '四、高可用与灾备'
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    ha_params = [
        ('需要高可用', 'TRUE/FALSE', '是否需要高可用架构'),
        ('可用性要求', '', '如：99.9%、99.99%、99.999%'),
        ('需要异地灾备', 'TRUE/FALSE', '是否需要异地灾备'),
        ('RPO (恢复点目标)', '', '可接受的数据丢失时间，如：0、5分钟、1小时'),
        ('RTO (恢复时间目标)', '', '可接受的系统恢复时间，如：5分钟、30分钟'),
        ('灾备中心数量', '', '需要的灾备中心数量'),
        ('主备切换方式', '', '自动/手动'),
    ]
    
    for param_name, default_value, description in ha_params:
        ws[f'A{row}'] = param_name
        ws[f'B{row}'] = default_value
        ws[f'C{row}'] = description
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        ws[f'C{row}'].border = thin_border
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].fill = input_fill
        row += 1
    
    row += 1
    
    # 第五部分：架构偏好
    ws[f'A{row}'] = '五、架构偏好'
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    arch_params = [
        ('需要读写分离', 'TRUE/FALSE', '是否需要读写分离'),
        ('需要分库分表', 'TRUE/FALSE', '是否需要分库分表'),
        ('分片策略', '', '如：按用户ID、按时间、按地域等'),
        ('需要缓存层', 'TRUE/FALSE', '是否需要Redis等缓存'),
        ('需要消息队列', 'TRUE/FALSE', '是否需要消息队列'),
        ('部署方式', '', '物理机/虚拟机/容器/云原生'),
        ('网络隔离要求', '', '是否需要VPC、专线等'),
    ]
    
    for param_name, default_value, description in arch_params:
        ws[f'A{row}'] = param_name
        ws[f'B{row}'] = default_value
        ws[f'C{row}'] = description
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        ws[f'C{row}'].border = thin_border
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].fill = input_fill
        row += 1
    
    row += 1
    
    # 第六部分：安全与合规
    ws[f'A{row}'] = '六、安全与合规'
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    security_params = [
        ('需要数据加密', 'TRUE/FALSE', '是否需要数据加密'),
        ('加密方式', '', '传输加密/存储加密/全加密'),
        ('需要审计日志', 'TRUE/FALSE', '是否需要完整的审计日志'),
        ('合规要求', '', '如：等保三级、PCI-DSS、GDPR等'),
        ('敏感数据类型', '', '如：身份证、银行卡、手机号等'),
        ('数据脱敏需求', 'TRUE/FALSE', '是否需要数据脱敏'),
    ]
    
    for param_name, default_value, description in security_params:
        ws[f'A{row}'] = param_name
        ws[f'B{row}'] = default_value
        ws[f'C{row}'] = description
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        ws[f'C{row}'].border = thin_border
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].fill = input_fill
        row += 1
    
    row += 1
    
    # 第七部分：运维需求
    ws[f'A{row}'] = '七、运维需求'
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].fill = section_fill
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    ops_params = [
        ('需要监控告警', 'TRUE/FALSE', '是否需要完整的监控告警'),
        ('备份频率', '', '如：实时、每小时、每天'),
        ('备份保留期限 (天)', '', '备份数据保留天数'),
        ('需要自动扩容', 'TRUE/FALSE', '是否需要自动扩容能力'),
        ('运维团队规模', '', '专职运维人员数量'),
        ('7x24支持需求', 'TRUE/FALSE', '是否需要7x24小时支持'),
    ]
    
    for param_name, default_value, description in ops_params:
        ws[f'A{row}'] = param_name
        ws[f'B{row}'] = default_value
        ws[f'C{row}'] = description
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        ws[f'C{row}'].border = thin_border
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].fill = input_fill
        row += 1
    
    wb.save('templates/部署资源预测模板-专业版.xlsx')
    print("✅ 专业版模板创建成功: templates/部署资源预测模板-专业版.xlsx")

if __name__ == '__main__':
    import os
    os.makedirs('templates', exist_ok=True)
    
    print("=" * 60)
    print("📝 创建Excel模板...")
    print("=" * 60)
    print()
    
    create_basic_template()
    create_professional_template()
    
    print()
    print("=" * 60)
    print("✅ 所有模板创建完成！")
    print("=" * 60)
    print()
    print("模板位置:")
    print("  - templates/部署资源预测模板-普通版.xlsx")
    print("  - templates/部署资源预测模板-专业版.xlsx")
