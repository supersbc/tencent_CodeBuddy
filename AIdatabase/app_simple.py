"""
TDSQL 部署资源预测系统 - 完整版
整合：部署预测、模型库管理、自主训练
"""

from flask import Flask, render_template, request, jsonify, redirect
import json
import os
from werkzeug.utils import secure_filename
from datetime import datetime

app = Flask(__name__)

# 配置
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'xlsx', 'xls', 'pdf', 'json', 'txt', 'log', 'gz', 'zip'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 2 * 1024 * 1024 * 1024  # 2GB max for log files

# 创建必要目录
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs('model_libraries', exist_ok=True)
os.makedirs('training_data', exist_ok=True)
os.makedirs('log_uploads', exist_ok=True)

print("\n" + "=" * 60)
print("🚀 TDSQL 部署资源预测系统 v4.2 (完整版)")
print("=" * 60)

# 延迟导入模块
predictor = None
library_manager = None
training_system = None
model = None

log_analyzer = None

def get_log_analyzer():
    """延迟加载日志分析器"""
    global log_analyzer
    if log_analyzer is None:
        print("📦 正在加载日志分析引擎...")
        from log_analyzer import TDSQLLogAnalyzer
        log_analyzer = TDSQLLogAnalyzer()
        print("✅ 日志分析引擎加载完成")
    return log_analyzer

def get_predictor():
    """延迟加载预测器"""
    global predictor
    if predictor is None:
        print("📦 正在加载预测引擎...")
        from deployment_predictor import DeploymentResourcePredictor
        predictor = DeploymentResourcePredictor()
        print("✅ 预测引擎加载完成")
    return predictor

def get_library_manager():
    """延迟加载模型库管理器"""
    global library_manager
    if library_manager is None:
        print("📦 正在加载模型库管理器...")
        from model_library_manager import ModelLibraryManager
        library_manager = ModelLibraryManager()
        print("✅ 模型库管理器加载完成")
    return library_manager

def get_training_system():
    """延迟加载训练系统"""
    global training_system, model
    if training_system is None:
        print("📦 正在加载训练系统...")
        from model import TDSQLArchitecturePredictor
        from training_system import TrainingSystem
        model = TDSQLArchitecturePredictor()
        training_system = TrainingSystem(model)
        print("✅ 训练系统加载完成")
    return training_system

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _normalize_vendor_list(value):
    if value is None:
        return []
    if isinstance(value, list):
        return [item for item in value if item]
    if isinstance(value, str):
        return [value] if value else []
    return []

# ==================== 全局错误处理（API返回JSON） ====================

@app.errorhandler(500)
def handle_500(e):
    import traceback
    traceback.print_exc()
    return jsonify({'error': f'服务器内部错误: {str(e)}'}), 500

@app.errorhandler(404)
def handle_404(e):
    if request.path.startswith('/api/'):
        return jsonify({'error': f'接口不存在: {request.path}'}), 404
    return e

# ==================== 页面路由 ====================

@app.route('/')
def index():
    """主页 - 默认进入导航页"""
    return redirect('/nav')

@app.route('/nav')
def navigation():
    """导航页面（别名）"""
    return render_template('navigation.html')

@app.route('/predict')
def predict_page():
    """部署资源预测页面（新版）"""
    return render_template('predict_v2.html')

@app.route('/predict_old')
def predict_page_old():
    """部署资源预测页面（旧版）"""
    return render_template('index.html')

@app.route('/model_library')
def model_library():
    """模型库管理页面"""
    return render_template('model_library.html')

@app.route('/learning')
def learning():
    """学习系统页面"""
    return render_template('index_learning.html')

@app.route('/log_analysis')
def log_analysis_page():
    """日志分析与健康评分页面"""
    return render_template('log_analysis.html')

@app.route('/experiment')
def experiment_page():
    """异常检测实验 - 技术方案与结果展示"""
    return render_template('experiment.html')

# ==================== API路由 ====================

@app.route('/api/health', methods=['GET'])
def health_check():
    """健康检查"""
    return jsonify({
        'status': 'ok',
        'version': '4.0',
        'message': 'TDSQL部署资源预测系统运行正常'
    })

@app.route('/api/predict', methods=['POST'])
def predict():
    """部署资源预测API"""
    try:
        raw = request.get_json() or {}
        
        # 统一字段映射（兼容普通版/专业版表单字段）
        data = {}
        
        # 数据规模
        data['data_volume'] = (
            raw.get('current_data_size_gb')
            or raw.get('total_data_size_gb')
            or raw.get('future_data_size_gb')
            or raw.get('data_volume')
            or 100
        )
        
        # 性能参数
        data['qps'] = raw.get('qps') or raw.get('normal_qps') or raw.get('peak_qps') or 1000
        data['tps'] = raw.get('tps') or raw.get('normal_tps') or raw.get('peak_tps') or int(data['qps'] * 0.3)
        
        # 并发/连接
        data['concurrent_users'] = (
            raw.get('concurrent_users')
            or raw.get('concurrent_connections')
            or raw.get('avg_concurrent_connections')
            or 100
        )
        
        # 行业/HA
        data['industry'] = raw.get('industry') or raw.get('industry_type') or 'general'
        need_ha = raw.get('need_high_availability') is True
        data['ha_level'] = 'high' if need_ha else 'standard'
        
        # 数据增长率（转为比例）
        growth = raw.get('data_growth_rate')
        if isinstance(growth, (int, float)) and growth > 1:
            data['data_growth_rate'] = float(growth) / 100.0
        else:
            data['data_growth_rate'] = growth if isinstance(growth, (int, float)) else 0.3
        
        # 其他可选需求
        data['need_disaster_recovery'] = bool(raw.get('need_disaster_recovery'))
        data['need_read_write_split'] = bool(raw.get('need_read_write_split'))

        # 厂商偏好（可多选）
        data['server_vendors'] = _normalize_vendor_list(raw.get('server_vendors'))
        data['network_vendors'] = _normalize_vendor_list(raw.get('network_vendors'))
        data['storage_vendors'] = _normalize_vendor_list(raw.get('storage_vendors'))
        data['software_vendors'] = _normalize_vendor_list(raw.get('software_vendors'))
        
        # 获取预测器
        pred = get_predictor()
        
        # 执行预测
        result = pred.predict(data)
        
        return jsonify({
            'success': True,
            'data': result
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """文件上传API"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '没有文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '文件名为空'}), 400
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{timestamp}_{filename}"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            # 获取预测器并解析文件
            pred = get_predictor()
            params = pred.parse_file(filepath)
            
            return jsonify({
                'success': True,
                'filename': filename,
                'filepath': filepath,
                'params': params
            })
        else:
            return jsonify({'success': False, 'error': '不支持的文件类型'}), 400
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/clear_uploads', methods=['POST'])
def clear_uploads():
    """清除上传的文件"""
    try:
        import shutil
        
        # 清空uploads目录
        if os.path.exists(UPLOAD_FOLDER):
            for filename in os.listdir(UPLOAD_FOLDER):
                file_path = os.path.join(UPLOAD_FOLDER, filename)
                try:
                    if os.path.isfile(file_path):
                        os.unlink(file_path)
                except Exception as e:
                    print(f"删除文件失败: {file_path}, 错误: {e}")
        
        return jsonify({
            'success': True,
            'message': '已清除所有上传文件'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== 模型库管理 API ====================

@app.route('/api/model_libraries', methods=['GET'])
def get_model_libraries():
    """获取可用的模型库列表"""
    try:
        manager = get_library_manager()
        libraries = manager.list_available_libraries()
        return jsonify({
            'success': True,
            'libraries': libraries
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/model_libraries/<library_id>', methods=['GET'])
def get_model_library(library_id):
    """获取指定模型库详情"""
    try:
        manager = get_library_manager()
        library = manager.get_library_info(library_id)
        return jsonify({
            'success': True,
            'library': library
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/model_libraries/<library_id>/download', methods=['POST'])
def download_model_library(library_id):
    """下载模型库"""
    try:
        manager = get_library_manager()
        result = manager.download_library(library_id)
        return jsonify({
            'success': True,
            'message': f'模型库 {library_id} 下载成功',
            'result': result
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/model_libraries/installed', methods=['GET'])
def get_installed_libraries():
    """获取已安装的模型库"""
    try:
        manager = get_library_manager()
        installed = manager.list_installed_libraries()
        return jsonify({
            'success': True,
            'libraries': installed
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/model_libraries/<library_id>/activate', methods=['POST'])
def activate_library(library_id):
    """激活模型库"""
    try:
        manager = get_library_manager()
        manager.activate_library(library_id)
        return jsonify({
            'success': True,
            'message': f'模型库 {library_id} 已激活'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/custom_library', methods=['POST'])
def create_custom_library():
    """创建自定义模型库"""
    try:
        data = request.get_json()
        manager = get_library_manager()
        result = manager.create_custom_library(
            name=data.get('name'),
            description=data.get('description'),
            industry=data.get('industry'),
            cases=data.get('cases', [])
        )
        return jsonify({
            'success': True,
            'library': result
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== 训练系统 API ====================

@app.route('/api/training/cases', methods=['GET'])
def get_training_cases():
    """获取训练案例列表"""
    try:
        trainer = get_training_system()
        cases = trainer.list_cases()
        return jsonify({
            'success': True,
            'cases': cases
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/training/cases', methods=['POST'])
def add_training_case():
    """添加训练案例"""
    try:
        data = request.get_json()
        trainer = get_training_system()
        
        case_id = trainer.add_case(
            input_data=data.get('input'),
            output_data=data.get('output'),
            feedback=data.get('feedback')
        )
        
        return jsonify({
            'success': True,
            'case_id': case_id,
            'message': '训练案例已添加'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/training/train', methods=['POST'])
def train_model():
    """训练模型"""
    try:
        data = request.get_json()
        trainer = get_training_system()
        
        result = trainer.train(
            epochs=data.get('epochs', 10),
            batch_size=data.get('batch_size', 32)
        )
        
        return jsonify({
            'success': True,
            'result': result
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/training/history', methods=['GET'])
def get_training_history():
    """获取训练历史"""
    try:
        trainer = get_training_system()
        history = trainer.get_history()
        return jsonify({
            'success': True,
            'history': history
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/training/evaluate', methods=['POST'])
def evaluate_model():
    """评估模型"""
    try:
        data = request.get_json()
        trainer = get_training_system()
        
        result = trainer.evaluate(data.get('test_data'))
        
        return jsonify({
            'success': True,
            'result': result
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== 日志分析 API ====================

@app.route('/api/log_analysis/analyze', methods=['POST'])
def analyze_logs():
    """日志分析API - 支持直接粘贴日志内容或上传文件"""
    try:
        analyzer = get_log_analyzer()

        # 获取参数
        window_size = 60
        window_step = 30

        # 方式1: 上传日志文件
        if 'log_file' in request.files:
            file = request.files['log_file']
            if file.filename:
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filepath = os.path.join('log_uploads', f"{timestamp}_{filename}")
                file.save(filepath)

                # 读取参数
                window_size = int(request.form.get('window_size', 60))
                window_step = int(request.form.get('window_step', 30))

                result = analyzer.analyze_file(filepath, window_size, window_step)
                return jsonify(result)

        # 方式2: JSON body（直接粘贴日志 或 指定日志文件路径）
        data = request.get_json() or {}

        window_size = int(data.get('window_size', window_size))
        window_step = int(data.get('window_step', window_step))

        # 从日志内容分析
        log_content = data.get('log_content', '')
        if log_content:
            result = analyzer.analyze(log_content, window_size, window_step)
            return jsonify(result)

        # 从日志文件路径分析
        log_path = data.get('log_path', '')
        if log_path:
            if not os.path.exists(log_path):
                return jsonify({'error': f'日志路径不存在: {log_path}'}), 400
            result = analyzer.analyze_file(log_path, window_size, window_step)
            return jsonify(result)

        return jsonify({'error': '请提供日志内容(log_content)、日志文件路径(log_path)或上传日志文件'}), 400

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/log_analysis/upload', methods=['POST'])
def upload_log_file():
    """上传日志文件"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '没有文件'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '文件名为空'}), 400

        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        save_name = f"{timestamp}_{filename}"
        filepath = os.path.join('log_uploads', save_name)
        file.save(filepath)

        # 读取前几行预览
        preview_lines = []
        try:
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                for i, line in enumerate(f):
                    if i >= 10:
                        break
                    preview_lines.append(line.rstrip())
        except Exception:
            pass

        # 统计文件行数
        line_count = 0
        try:
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                for _ in f:
                    line_count += 1
        except Exception:
            pass

        return jsonify({
            'success': True,
            'filename': save_name,
            'filepath': filepath,
            'line_count': line_count,
            'preview': preview_lines,
            'file_size': os.path.getsize(filepath)
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/log_analysis/browse', methods=['POST'])
def browse_directory():
    """浏览服务器目录，选择日志文件"""
    try:
        data = request.get_json() or {}
        path = data.get('path', '/data/workspace')

        if not os.path.exists(path):
            return jsonify({'error': f'路径不存在: {path}'}), 400

        if os.path.isfile(path):
            # 返回文件信息
            size = os.path.getsize(path)
            preview = []
            try:
                with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                    for i, line in enumerate(f):
                        if i >= 5:
                            break
                        preview.append(line.rstrip()[:200])
            except Exception:
                pass
            return jsonify({
                'type': 'file',
                'path': path,
                'name': os.path.basename(path),
                'size': size,
                'size_mb': round(size / 1024 / 1024, 1),
                'preview': preview
            })

        # 列出目录内容
        items = []
        try:
            entries = sorted(os.listdir(path))
        except PermissionError:
            return jsonify({'error': '无权限访问该目录'}), 403

        for name in entries:
            if name.startswith('.'):
                continue
            full = os.path.join(path, name)
            try:
                if os.path.isdir(full):
                    items.append({'name': name, 'type': 'dir', 'path': full})
                else:
                    size = os.path.getsize(full)
                    ext = os.path.splitext(name)[1].lower()
                    # 只显示可能是日志的文件
                    is_log = ext in ('.log', '.txt', '.json', '.gz', '.zip', '') or 'log' in name.lower() or 'interf' in name.lower()
                    if is_log or size < 10 * 1024:  # 小文件也显示
                        items.append({
                            'name': name,
                            'type': 'file',
                            'path': full,
                            'size': size,
                            'size_mb': round(size / 1024 / 1024, 1)
                        })
            except Exception:
                continue

        return jsonify({
            'type': 'dir',
            'path': path,
            'parent': os.path.dirname(path),
            'items': items
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/log_analysis/chunk_upload', methods=['POST'])
def chunk_upload():
    """分片上传大文件"""
    try:
        chunk = request.files.get('chunk')
        if not chunk:
            return jsonify({'success': False, 'error': '无分片数据'}), 400

        filename = request.form.get('filename', 'unknown')
        chunk_index = int(request.form.get('chunk_index', 0))
        total_chunks = int(request.form.get('total_chunks', 1))
        upload_id = request.form.get('upload_id', '')

        safe_name = secure_filename(filename)
        if not upload_id:
            upload_id = datetime.now().strftime('%Y%m%d_%H%M%S')

        target_path = os.path.join('log_uploads', f"{upload_id}_{safe_name}")

        mode = 'ab' if chunk_index > 0 else 'wb'
        with open(target_path, mode) as f:
            f.write(chunk.read())

        # 最后一个分片
        if chunk_index == total_chunks - 1:
            final_size = os.path.getsize(target_path)
            preview = []
            try:
                with open(target_path, 'r', encoding='utf-8', errors='ignore') as f:
                    for i, line in enumerate(f):
                        if i >= 5:
                            break
                        preview.append(line.rstrip()[:200])
            except Exception:
                pass
            return jsonify({
                'success': True,
                'done': True,
                'filepath': target_path,
                'filename': f"{upload_id}_{safe_name}",
                'file_size': final_size,
                'preview': preview
            })

        return jsonify({
            'success': True,
            'done': False,
            'chunk_index': chunk_index,
            'total_chunks': total_chunks,
            'progress': round((chunk_index + 1) / total_chunks * 100, 1)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/download_template/<version>')
def download_template(version):
    """下载Excel模板"""
    try:
        from flask import send_file
        if version == 'basic':
            filename = '部署资源预测模板-普通版.xlsx'
        else:
            filename = '部署资源预测模板-专业版.xlsx'
        
        filepath = os.path.join('templates', filename)
        return send_file(filepath, as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/parse_excel', methods=['POST'])
def parse_excel():
    """解析上传的Excel文件"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '没有文件上传'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '文件名为空'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': '只支持Excel文件'}), 400
        
        # 保存文件
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # 解析Excel
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        ws = wb.active
        
        data = {}
        
        # 读取数据（从第5行开始，跳过标题和表头）
        for row in range(5, ws.max_row + 1):
            param_name = ws.cell(row, 1).value
            param_value = ws.cell(row, 2).value
            
            if param_name and param_value:
                # 转换参数名为字段名
                field_mapping = {
                    '数据规模 (GB)': 'total_data_size_gb',
                    '当前数据规模 (GB)': 'current_data_size_gb',
                    '表数量': 'table_count',
                    'QPS (每秒查询数)': 'qps',
                    '日常QPS': 'normal_qps',
                    'TPS (每秒事务数)': 'tps',
                    '日常TPS': 'normal_tps',
                    '并发连接数': 'concurrent_connections',
                    '平均并发连接数': 'avg_concurrent_connections',
                    '需要高可用': 'need_high_availability',
                    '需要灾备': 'need_disaster_recovery',
                    '需要异地灾备': 'need_disaster_recovery',
                    '需要读写分离': 'need_read_write_split',
                    '数据增长率 (%/年)': 'data_growth_rate',
                }
                
                field_name = field_mapping.get(param_name)
                if field_name:
                    # 处理布尔值
                    if isinstance(param_value, str) and param_value.upper() in ['TRUE', 'FALSE']:
                        data[field_name] = param_value.upper() == 'TRUE'
                    else:
                        data[field_name] = param_value
        
        # 删除临时文件
        os.remove(filepath)
        
        return jsonify({
            'success': True,
            'data': data
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    print("\\n" + "=" * 60)
    print("🚀 TDSQL 部署资源预测系统 v4.3")
    print("=" * 60)
    print(f"📍 主页面: http://127.0.0.1:18080")
    print(f"📍 导航页面: http://127.0.0.1:18080/nav")
    print(f"📍 部署预测: http://127.0.0.1:18080/predict (新版)")
    print(f"📍 模型库管理: http://127.0.0.1:18080/model_library")
    print(f"📍 学习系统: http://127.0.0.1:18080/learning")
    print(f"📍 日志分析: http://127.0.0.1:18080/log_analysis")
    print(f"📍 实验分析: http://127.0.0.1:18080/experiment")
    print("=" * 60)
    print("✨ 功能模块:")
    print("  ✅ 部署资源预测 - 普通版/专业版双模式")
    print("  ✅ Excel模板 - 下载模板，上传自动填充")
    print("  ✅ 模型库管理 - 8个预置模型库")
    print("  ✅ 自主训练 - 从实际案例中学习优化")
    print("  ✅ 文件上传 - 支持JSON/Excel/图片/PDF")
    print("  ✅ 日志分析 - 基于E-Log的数据库健康评分")
    print("  ✅ 实验分析 - 异常检测技术方案与结果展示")
    print("=" * 60)
    print()
    
    app.run(
        host='0.0.0.0',
        port=18080,
        debug=False,
        threaded=True,
        use_reloader=False
    )
